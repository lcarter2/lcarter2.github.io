import os
from openpyxl import load_workbook

def update_files(directory):

    os.chdir(directory)
    directory_completed = os.fsencode(directory)
    for file in os.listdir(directory_completed):
        filename = os.fsdecode(file)
        if '.xlsx' in filename and filename[:2] != '~$': #Avoid temporary files
            wb = load_workbook(filename)
            ws = wb['Form']
            
            old_num = ws['C5'].value
            new_num = filename.split('_')[0]
            
            if int(old_num) != int(new_num):
                print(old_num, " -> ", new_num)
            
            ws.cell(row=5, column=3).value = int(new_num)
            
            wb.save(filename)
            wb.close()
        
        
def main():
    #update_files('C:\\ForLogan\\MSTestCases\\Foo')
    filename = '100_sally.xlsx'
    wb = load_workbook(filename)
    ws = wb['Form']
    
    old_num = ws['C5'].value
    new_num = filename.split('_')[0]
    
    if int(old_num) != int(new_num):
        print(old_num, " -> ", new_num)
    
    ws.cell(row=5, column=3).value = int(new_num)
    
    wb.save(filename)
    wb.close()